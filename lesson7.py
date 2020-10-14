'''
openpyxl:第三方库 实现对excel文件读取及编辑
1、安装openpyxl   pip install openpyxl

2、openpyxl应用
requests包含三大组件：
1、工作簿
2、工作表
3、单元格
'''
import openpyxl

def ReadData(filename,sheetname):
    # 获取工作簿
    wk=openpyxl.open(filename=filename)
    # 获取当前工作表
    sheet1=wk[sheetname]
    # print(sheet1.cell(row=1,column=1).value)
    # 获取最大行数
    # print('获取最大行数：',sheet1.max_row)
    # 获取最大列数
    # print('获取最大列数:',sheet1.max_column)

    # # 如何取得excel所有的数据\
    datalist=[]
    for x in range(2,15):
        # 把用例数据存储到字典
        # print(sheet1.cell(row=x, column=1).value)
        case=dict(url=sheet1.cell(row=x,column=5).value,
             data=sheet1.cell(row=x,column=6).value,
             expected=sheet1.cell(row=x,column=7).value)
        # print(case)
        datalist.append(case)
        #是否需要返回值
    return datalist

# 取excel数据
# list1=ReadData(filename='test_case_api.xlsx',sheetname='login')
# print(list1)

# openpyx模块编辑excel文件数据
wk1=openpyxl.open('test_case_api.xlsx')
register=wk1["register"]
register.cell(2,8).value='ok'
wk1.save('test_case_api.xlsx')  #保存，更新的内容才生效






